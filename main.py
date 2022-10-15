import gi
gi.require_version("Gtk", "3.0")
from gi.repository import Gtk
from openpyxl import Workbook
from openpyxl.chart import ( LineChart, Reference )
from openpyxl.chart.axis import DateAxis
import random

def make_xsls():
    values_count = 70

    wb = Workbook()
    worksheet = wb.create_sheet("Hello", 0)


    worksheet['A1'] = "Date"
    worksheet['B1'] = "Value"

    for i in range(values_count):
        worksheet.cell(row=2+i, column=1, value=i)

    for i in range(values_count):
        worksheet.cell(row=2+i, column=2, value=random.randrange(10,100))

    chart = LineChart()
    chart.title = "Values over time"
    chart.style = 1
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Day No"
    chart.width = 50
    chart.height= 10
    data = Reference(worksheet, min_col=2, min_row=2, max_col=2, max_row=1+values_count)
    chart.add_data(data, titles_from_data=False)
    worksheet.add_chart(chart, "D3")

    wb.save('form1.xlsx')

# Sa fac ca aici:
# https://python-gtk-3-tutorial.readthedocs.io/en/latest/layout.html#id3

class MyWindow(Gtk.Window):
    def __init__(self):
        super().__init__(title="Hello")

        box_v = Gtk.Box(orientation="vertical", spacing=0, margin=20)
        self.add(box_v)
        

        # Rand 1 - Category 1
        # checkbox si label
        ck_cat_1 = Gtk.CheckButton()
        label_cat_1 = Gtk.Label( label="Category 1", halign=Gtk.Align.START )


        # Rand 2 - Sub-Category 1-1
        # checkbox si label
        ck_subcat_1_1 = Gtk.CheckButton()
        label_subcat_1 = Gtk.Label( label="Subcategory 1-1", halign=Gtk.Align.START )


        grid = Gtk.Grid()
        grid.attach(ck_cat_1, 0, 0, 1, 1)
        grid.attach(label_cat_1, 2, 0, 1, 1)
        grid.attach(ck_subcat_1_1, 1, 1, 1, 1)
        grid.attach(label_subcat_1, 2, 1, 1, 1)
        box_v.pack_start(grid, True, True, 0)


        self.button = Gtk.Button(label="Generate", margin=20, margin_top=40)
        self.button.connect("clicked", self.on_button_clicked)
        box_v.pack_start(self.button, True, True, 0)


    def on_button_clicked(self, widget):
        print("to do")


win = MyWindow()
win.connect("destroy", Gtk.main_quit)
win.show_all()
Gtk.main()
