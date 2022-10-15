from openpyxl import Workbook
from openpyxl.chart import ( LineChart, Reference )
from openpyxl.chart.axis import DateAxis
import random

# categories - list of string.
# Represents the name of the columns to be added
def make_form_A(categories):
    values_count = 70

    wb = Workbook()
    worksheet = wb.create_sheet("Form A", 0)

    # Put categories
    for index, c in enumerate(categories):
        worksheet.cell(row=1, column=1+index, value=c)

    for i in range(values_count):
        worksheet.cell(row=2+i, column=1, value=i)

    for i in range(values_count):
        worksheet.cell(row=2+i, column=2, value=random.randrange(10,100))

    chart = LineChart()
    chart.title = "Values over time"
    chart.style = 1
    chart.y_axis.title = "Value"
    chart.x_axis.title = categories[0]
    chart.width = 30
    chart.height= 10
    data = Reference(worksheet, min_col=2, min_row=2, max_col=2, max_row=1+values_count)
    chart.add_data(data, titles_from_data=False)
    worksheet.add_chart(chart, "D3")
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 18

    wb.save('formA.xlsx')

def make_form_B(categories):
    values_count = 70

    wb = Workbook()
    worksheet = wb.create_sheet("Form B", 0)

    # Put categories
    for index, c in enumerate(categories):
        worksheet.cell(row=1, column=1+index, value=c)

    for i in range(values_count):
        worksheet.cell(row=2+i, column=1, value=i)

    for i in range(values_count):
        worksheet.cell(row=2+i, column=2, value=random.randrange(10,100))

    chart = LineChart()
    chart.title = "Values over time"
    chart.style = 1
    chart.y_axis.title = "Value"
    chart.x_axis.title = categories[0]
    chart.width = 50
    chart.height= 10
    data = Reference(worksheet, min_col=2, min_row=2, max_col=2, max_row=1+values_count)
    chart.add_data(data, titles_from_data=False)
    worksheet.add_chart(chart, "D3")

    wb.save('formB.xlsx')

def make_form(form):
    categories_to_add = []
    for index, c in enumerate(form.subcategories_checkbox):
        if c.get_active() == True:
            categories_to_add.append(form.subcategories[index])

    if form.name == "A":
        make_form_A(categories_to_add)

    elif form.name == "B":
        make_form_B(categories_to_add)
